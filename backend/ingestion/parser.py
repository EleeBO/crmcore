"""File parsers for PDF, Excel, DOCX, Markdown, plain text."""

from __future__ import annotations

from dataclasses import dataclass, field

from backend.errors import ErrorCode, IngestionError
from backend.logger import logger

SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".docx", ".md", ".txt"}


@dataclass
class ParsedChunk:
    text: str
    source_file: str = ""
    page_number: int = 1
    section_title: str = ""
    chunk_type: str = "text"  # "text" | "table"
    metadata: dict = field(default_factory=dict)


def parse_file(data: bytes, filename: str) -> list[ParsedChunk]:
    """Dispatch to the correct parser based on file extension."""
    ext = _ext(filename)
    if ext not in SUPPORTED_EXTENSIONS:
        raise IngestionError(
            ErrorCode.FILE_UNSUPPORTED,
            f"Unsupported file type: {ext!r}. Supported: {sorted(SUPPORTED_EXTENSIONS)}",  # noqa: E501
        )
    if ext == ".pdf":
        return parse_pdf(data, source_file=filename)
    if ext in (".xlsx", ".xls"):
        return parse_excel(data, source_file=filename)
    if ext == ".docx":
        return parse_docx(data, source_file=filename)
    # .md / .txt
    return _parse_text(data.decode("utf-8", errors="replace"), source_file=filename)


def parse_pdf(data: bytes, source_file: str = "") -> list[ParsedChunk]:
    """Extract text from PDF pages using PyMuPDF."""
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:
        raise IngestionError(ErrorCode.FILE_CORRUPT, "PyMuPDF not installed") from exc

    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception as exc:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            f"Cannot open PDF: {exc}",
            detail=str(exc),
        ) from exc

    chunks: list[ParsedChunk] = []
    for page_idx in range(len(doc)):
        page = doc[page_idx]
        text = page.get_text().strip()
        if not text:
            continue
        chunks.append(
            ParsedChunk(
                text=text,
                source_file=source_file,
                page_number=page_idx + 1,
                chunk_type="text",
            )
        )

    if not chunks:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            "PDF contains no extractable text (may be scanned or encrypted).",
        )

    logger.debug(f"Parsed PDF {source_file!r}: {len(chunks)} page-chunks")
    return chunks


def parse_excel(data: bytes, source_file: str = "") -> list[ParsedChunk]:
    """Parse XLSX/XLS into row-based chunks with column headers preserved."""
    import io

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngestionError(ErrorCode.FILE_CORRUPT, "openpyxl not installed") from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            f"Cannot open Excel file: {exc}",
            detail=str(exc),
        ) from exc

    chunks: list[ParsedChunk] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]
        for row_idx, row in enumerate(rows[1:], start=2):
            parts = []
            for header, value in zip(headers, row, strict=False):
                if value is not None and str(value).strip():
                    parts.append(f"{header}: {value}")
            if not parts:
                continue
            text = " | ".join(parts)
            chunks.append(
                ParsedChunk(
                    text=text,
                    source_file=source_file,
                    page_number=row_idx,
                    section_title=sheet.title,
                    chunk_type="table",
                    metadata={"sheet": sheet.title, "row": row_idx},
                )
            )

    if not chunks:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            "Excel file contains no data rows.",
        )

    logger.debug(f"Parsed Excel {source_file!r}: {len(chunks)} row-chunks")
    return chunks


def parse_docx(data: bytes, source_file: str = "") -> list[ParsedChunk]:
    """Parse DOCX into text chunks, capturing section headings."""
    import io

    try:
        from docx import Document
    except ImportError as exc:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT, "python-docx not installed"
        ) from exc

    try:
        doc = Document(io.BytesIO(data))
    except Exception as exc:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            f"Cannot open DOCX file: {exc}",
            detail=str(exc),
        ) from exc

    chunks: list[ParsedChunk] = []
    current_section = ""
    current_paragraphs: list[str] = []

    def _flush() -> None:
        if current_paragraphs:
            text = "\n".join(current_paragraphs).strip()
            if text:
                chunks.append(
                    ParsedChunk(
                        text=text,
                        source_file=source_file,
                        page_number=len(chunks) + 1,
                        section_title=current_section,
                        chunk_type="text",
                    )
                )

    for para in doc.paragraphs:
        if para.style.name.startswith("Heading"):
            _flush()
            current_section = para.text.strip()
            current_paragraphs = []
        elif para.text.strip():
            current_paragraphs.append(para.text.strip())

    _flush()

    if not chunks:
        raise IngestionError(
            ErrorCode.FILE_CORRUPT,
            "DOCX contains no extractable text.",
        )

    logger.debug(f"Parsed DOCX {source_file!r}: {len(chunks)} section-chunks")
    return chunks


def _parse_text(text: str, source_file: str = "") -> list[ParsedChunk]:
    """Parse plain text or Markdown."""
    return [ParsedChunk(text=text, source_file=source_file, page_number=1)]


def _ext(filename: str) -> str:
    import os

    return os.path.splitext(filename.lower())[1]
